"""
Génère le fichier .xlsx pushable dans DECA depuis les décisions VALIDÉ.

Le fichier produit contient une ligne par DECA décidé, avec les colonnes
attendues par DECA pour l'import en masse (même ordre que l'extract source).

Usage:
    python -m scripts.export_deca                   # tous modules
    python -m scripts.export_deca --module SM21     # un module
    python -m scripts.export_deca --out exports/    # dossier de sortie
"""
import sys
import argparse
import logging
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import EXPORTS_DIR, MODULES
from db.db import init_schema
from db import queries

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)


# ── Colonnes de l'export — ordre attendu par DECA ────────────────────────────
# Les noms correspondent aux champs DECA (format may_full)

EXPORT_COLUMNS = [
    # Identifiants
    ("marquage",            "marquage"),
    ("ref_constructeur",    "ref_constructeur"),
    # Infos outil (lecture seule dans DECA, on les reprend telles quelles)
    ("etat",                "etat"),
    ("famille",             "famille"),
    ("sous_famille",        "sous_famille"),
    ("type_outil",          "type_outil"),
    ("constructeur",        "constructeur"),
    ("nserie",              "nserie"),
    # Services — nouvelles valeurs décidées
    ("n_service1",          "service1"),
    ("n_service2",          "service2"),
    ("n_service3",          "service3"),
    ("n_service4",          "service4"),
    # Localisation — reprise telle quelle
    ("localisation1",       "localisation1"),
    ("localisation2",       "localisation2"),
    ("localisation3",       "localisation3"),
    # Meta décision
    ("decision",            "decision"),
    ("pre_check",           "pre_check"),
    ("commentaire",         "commentaire"),
    ("updated_at",          "updated_at"),
    ("module_context",      "module"),
]

# Nom de colonne dans l'export final → label affiché dans le fichier
HEADER_LABELS = {
    "marquage":         "Marquage",
    "ref_constructeur": "Réf. Constructeur",
    "etat":             "Etat",
    "famille":          "Famille",
    "sous_famille":     "Sous-famille",
    "type_outil":       "Type",
    "constructeur":     "Constructeur",
    "nserie":           "N Série",
    "service1":         "Service1",
    "service2":         "Service2",
    "service3":         "Service3",
    "service4":         "Service4",
    "localisation1":    "Localisation1",
    "localisation2":    "Localisation2",
    "localisation3":    "Localisation3",
    "decision":         "Décision",
    "pre_check":        "Pré-check",
    "commentaire":      "Commentaire",
    "updated_at":       "Mis à jour le",
    "module":           "Module",
}


def build_export_df(module: str | None = None) -> pd.DataFrame:
    """
    Construit le DataFrame export depuis les décisions VALIDÉ.
    Chaque ligne = 1 DECA avec les nouveaux services décidés.
    """
    rows = queries.get_decisions_for_export(module)
    if not rows:
        return pd.DataFrame()

    records = []
    for r in rows:
        record = {}
        for src_field, out_field in EXPORT_COLUMNS:
            record[out_field] = r[src_field] if src_field in r.keys() else ""
        records.append(record)

    df = pd.DataFrame(records)

    # Formater la date updated_at
    if "updated_at" in df.columns:
        df["updated_at"] = pd.to_datetime(
            df["updated_at"], utc=True, errors="coerce"
        ).dt.strftime("%d/%m/%Y %H:%M")

    df.rename(columns=HEADER_LABELS, inplace=True)
    return df


def export(module: str | None = None, out_dir: Path | None = None) -> Path:
    """
    Génère le fichier .xlsx et le sauvegarde dans out_dir.
    Retourne le chemin du fichier généré.
    """
    out_dir = out_dir or EXPORTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    df = build_export_df(module)
    if df.empty:
        raise ValueError(
            f"Aucune décision VALIDÉ à exporter"
            + (f" pour le module {module}" if module else "")
        )

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    suffix = f"_{module}" if module else "_ALL"
    filename = f"export_DECA{suffix}_{ts}.xlsx"
    out_path = out_dir / filename

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Décisions VALIDÉ")

        # Mise en forme basique
        ws = writer.sheets["Décisions VALIDÉ"]
        _format_sheet(ws, df)

    log.info("Export généré : %s (%d lignes)", out_path, len(df))
    return out_path


def _format_sheet(ws, df: pd.DataFrame):
    """Applique une mise en forme minimale : largeurs + header gras."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    col_widths = {
        "Marquage": 10,
        "Réf. Constructeur": 18,
        "Etat": 12,
        "Famille": 14,
        "Sous-famille": 14,
        "Type": 14,
        "Constructeur": 14,
        "N Série": 12,
        "Service1": 10,
        "Service2": 10,
        "Service3": 12,
        "Service4": 18,
        "Localisation1": 14,
        "Localisation2": 14,
        "Localisation3": 14,
        "Décision": 12,
        "Pré-check": 10,
        "Commentaire": 30,
        "Mis à jour le": 16,
        "Module": 8,
    }

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

        letter = get_column_letter(col_idx)
        width = col_widths.get(col_name, 14)
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export décisions VALIDÉ vers .xlsx")
    parser.add_argument("--module", choices=MODULES, default=None,
                        help="Filtrer sur un module (défaut : tous)")
    parser.add_argument("--out", type=Path, default=None,
                        help="Dossier de sortie (défaut : exports/)")
    args = parser.parse_args()

    init_schema()
    try:
        path = export(module=args.module, out_dir=args.out)
        print(f"Fichier généré : {path}")
    except ValueError as e:
        print(f"Erreur : {e}")
        sys.exit(1)
